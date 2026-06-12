"""
DF REG MKT Teammate — Site Generator
Usage: python generate_site.py
Reads the Excel file and outputs index.html in the same folder.
Re-run this script whenever the Excel content is updated.
"""
from openpyxl import load_workbook
import json, os, re

EXCEL = r"C:\Users\lintim\Desktop\claude\DF REG MKT\DF REG MKT Teammate\DF Regional Marketing TeamMate.xlsx"
HERE  = os.path.dirname(os.path.abspath(__file__))

def c(v):
    return str(v).strip() if v is not None else ""

# ─── Parse Excel ──────────────────────────────────────────────────────────────
wb = load_workbook(EXCEL, data_only=True)
data = {}

# Standard tabs: row1=title, row2=headers, row3+=data
STD = ["scenario","pic","steps","prepare","timeline","qa","links"]
for name in wb.sheetnames:
    if name in ('Index','Reg & Local Contactor','Social Media Link'):
        continue
    ws = wb[name]
    rows = []
    for r in ws.iter_rows(min_row=3):
        vals = [cell.value for cell in r[:7]]
        if any(v is not None for v in vals):
            row = dict(zip(STD, [c(v) for v in vals]))
            # Read hyperlinks from all columns and store as col_url
            for i, cell in enumerate(r[:7]):
                if cell.hyperlink and cell.hyperlink.target:
                    row[STD[i] + '_url'] = cell.hyperlink.target
            rows.append(row)
    data[name] = rows

# Index tab
ws = wb['Index']
idx = []
for r in ws.iter_rows(min_row=3, values_only=True):
    if any(c(v) for v in r):
        idx.append({"tab": c(r[0]), "name": c(r[1]), "desc": c(r[2])})
data['_index'] = idx

# Contactor tab
ws = wb['Reg & Local Contactor']
team, contactors, sec = [], [], None
for r in ws.iter_rows(values_only=True):
    if not any(v for v in r): continue
    f = c(r[0])
    if 'Regional Team' in f:  sec = None; continue
    if f == 'Member':         sec = 'team'; continue
    if 'Local MKT' in f:     sec = None; continue
    if f == 'Region':         sec = 'loc'; continue
    if sec == 'team':
        team.append({"name": c(r[0]), "resp": c(r[1]), "email": c(r[2])})
    elif sec == 'loc':
        contactors.append({"region": c(r[0]), "contact": c(r[1]), "email": c(r[2])})
data['_contactor'] = {"team": team, "contactors": contactors}

# Social Media Links
ws   = wb['Social Media Link']
sm   = {"off_hdr":[], "off":[], "non_hdr":[], "non":[], "esp_hdr":[], "esp":[]}
sec  = None
for r in ws.iter_rows(values_only=True):
    first = c(r[0])
    if first == 'Official Channels': sec='off';  continue
    if first == 'Non-Official':      sec='non';  continue
    if first == 'Esports Channel':   sec='esp';  continue
    if not any(v for v in r):        continue

    if sec == 'off':
        if first == 'Channel':
            sm['off_hdr'] = [c(v) for v in r if c(v)]
        else:
            sm['off'].append([c(v) if c(v) else '—' for v in r[:len(sm['off_hdr'])]])
    elif sec == 'non':
        if first == 'Channel':
            sm['non_hdr'] = [c(v) for v in r if c(v)]
        else:
            sm['non'].append([c(v) if c(v) else '—' for v in r[:len(sm['non_hdr'])]])
    elif sec == 'esp':
        if first == 'Channel':
            sm['esp_hdr'] = [c(v) for v in r if c(v)]
        else:
            row_vals = [c(v) if c(v) else '—' for v in r]
            while row_vals and row_vals[-1] == '—': row_vals.pop()
            if any(v != '—' for v in row_vals):
                sm['esp'].append(row_vals)
data['_sm'] = sm

# ─── Nav definition ───────────────────────────────────────────────────────────
NAV = [
    {"id":"_index",    "label":"Index",              "type":"index"},
    {"id":"_contactor","label":"Contactor",           "type":"contactor"},
    {"id":"_sm",       "label":"SM Links",            "type":"sm"},
    {"id":"sep1",      "sep": True},
    {"id":"tab1",  "label":"Tab 1 · Budget & Plan",  "sheet":"Tab1 Budget Plan & Report"},
    {"id":"tab2",  "label":"Tab 2 · OM",             "sheet":"Tab2 OM"},
    {"id":"tab3",  "label":"Tab 3 · Assets",         "sheet":"Tab3 Assets"},
    {"id":"tab4",  "label":"Tab 4 · Social Media",   "sheet":"Tab4 Social Media"},
    {"id":"tab5",  "label":"Tab 5 · PR",             "sheet":"Tab5 PR"},
    {"id":"tab7",  "label":"Tab 7 · Esports",        "sheet":"Tab7 Esports"},
    {"id":"tab8",  "label":"Tab 8 · Partnership & IP","sheet":"Tab8 Partnership & IP"},
    {"id":"tab9",  "label":"Tab 9 · In-game Items",  "sheet":"Tab9 In-game Items Requirement"},
    {"id":"tab10", "label":"Tab 10 · CCP",           "sheet":"Tab10 CCP"},
    {"id":"tab11", "label":"Tab 11 · Website",       "sheet":"Tab11 Website"},
    {"id":"tab12", "label":"Tab 12 · Patch Updates", "sheet":"Tab12 Patch Updates"},
    {"id":"tab13", "label":"Tab 13 · Store Feature", "sheet":"Tab13 Store Feature"},
]

tabs = {}
tabs['_index']     = data['_index']
tabs['_contactor'] = data['_contactor']
tabs['_sm']        = data['_sm']
for item in NAV:
    if 'sheet' in item:
        tabs[item['id']] = data.get(item['sheet'], [])

DATA_JSON = json.dumps({"nav": NAV, "tabs": tabs}, ensure_ascii=False, indent=None)

# ─── HTML template ────────────────────────────────────────────────────────────
HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>DF REG MKT Teammate</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;font-size:14px;background:#f1f5f9;color:#1e293b;display:flex;height:100vh;overflow:hidden}

/* ── Sidebar ── */
#sidebar{width:228px;min-width:228px;background:#0f172a;color:#94a3b8;display:flex;flex-direction:column;overflow-y:auto;flex-shrink:0}
#sidebar-header{padding:20px 16px 14px;border-bottom:1px solid #1e293b}
#sidebar-header .logo{font-size:11px;font-weight:700;letter-spacing:.1em;color:#64748b;text-transform:uppercase;margin-bottom:4px}
#sidebar-header .title{font-size:15px;font-weight:700;color:#f1f5f9;line-height:1.3}
#nav-list{list-style:none;padding:8px 0;flex:1}
#nav-list li.sep{height:1px;background:#1e293b;margin:8px 12px}
#nav-list li.nav-item a{display:block;padding:8px 16px;color:#94a3b8;text-decoration:none;font-size:13px;border-radius:0;transition:background .15s,color .15s;line-height:1.4}
#nav-list li.nav-item a:hover{background:#1e293b;color:#e2e8f0}
#nav-list li.nav-item a.active{background:#1d4ed8;color:#fff;font-weight:600}

/* ── Main ── */
#main{flex:1;display:flex;flex-direction:column;overflow:hidden}
#toolbar{padding:14px 24px;background:#fff;border-bottom:1px solid #e2e8f0;display:flex;align-items:center;gap:12px;flex-shrink:0}
#toolbar h2{font-size:16px;font-weight:700;color:#0f172a;min-width:0;flex:1;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
#search{padding:7px 12px;border:1px solid #cbd5e1;border-radius:6px;font-size:13px;width:260px;outline:none;background:#f8fafc}
#search:focus{border-color:#3b82f6;background:#fff}
#content{flex:1;overflow-y:auto;padding:20px 24px}

/* ── Cards ── */
.cards-grid{display:flex;flex-direction:column;gap:12px}
.card{background:#fff;border:1px solid #e2e8f0;border-radius:10px;overflow:hidden;transition:box-shadow .15s}
.card:hover{box-shadow:0 2px 12px rgba(0,0,0,.07)}
.card-header{padding:14px 16px;cursor:pointer;display:flex;align-items:flex-start;gap:10px;user-select:none}
.card-header:hover{background:#f8fafc}
.card-scenario{flex:1;font-weight:600;font-size:14px;color:#0f172a;line-height:1.4;white-space:pre-line}
.card-pic{flex-shrink:0;padding:2px 9px;border-radius:20px;font-size:11px;font-weight:700;letter-spacing:.03em;margin-top:2px}
.card-chevron{color:#94a3b8;font-size:12px;margin-top:3px;flex-shrink:0;transition:transform .2s}
.card-chevron.open{transform:rotate(180deg)}
.card-body{display:none;padding:0 16px 16px;border-top:1px solid #f1f5f9}
.card-body.open{display:block}
.field{margin-top:14px}
.field-label{font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.07em;color:#64748b;margin-bottom:5px}
.field-value{color:#1e293b;line-height:1.7;white-space:pre-wrap;word-break:break-word;font-size:13.5px}
.field-value a{color:#2563eb;text-decoration:none}
.field-value a:hover{text-decoration:underline}
.field-link{display:inline-block;margin-top:6px;padding:4px 10px;background:#eff6ff;border:1px solid #bfdbfe;border-radius:5px;color:#1d4ed8;font-size:12px;font-weight:600;text-decoration:none}
.field-link:hover{background:#dbeafe}

/* PIC colors */
.pic-tim{background:#ede9fe;color:#5b21b6}
.pic-ruru{background:#fce7f3;color:#9d174d}
.pic-jeremy{background:#fef3c7;color:#92400e}
.pic-hua{background:#d1fae5;color:#065f46}
.pic-mye{background:#dbeafe;color:#1e40af}
.pic-multi{background:#f1f5f9;color:#475569}

/* ── Tables (Index / Contactor / SM) ── */
.page-title{font-size:18px;font-weight:700;color:#0f172a;margin-bottom:18px}
.section-title{font-size:13px;font-weight:700;text-transform:uppercase;letter-spacing:.07em;color:#64748b;margin:24px 0 10px}
.data-table{width:100%;border-collapse:collapse;background:#fff;border:1px solid #e2e8f0;border-radius:10px;overflow:hidden}
.data-table th{background:#f8fafc;padding:10px 14px;text-align:left;font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:.05em;color:#64748b;border-bottom:1px solid #e2e8f0;white-space:nowrap}
.data-table td{padding:10px 14px;border-bottom:1px solid #f1f5f9;vertical-align:top;font-size:13px;line-height:1.6;word-break:break-word}
.data-table tr:last-child td{border-bottom:none}
.data-table td a{color:#2563eb;text-decoration:none}
.data-table td a:hover{text-decoration:underline}
.data-table .resp-col{white-space:pre-wrap;max-width:360px}
.sm-table-wrap{overflow-x:auto;border-radius:10px;border:1px solid #e2e8f0;margin-bottom:20px}
.sm-table{border-collapse:collapse;background:#fff;min-width:600px}
.sm-table th{background:#f8fafc;padding:9px 12px;text-align:left;font-size:12px;font-weight:700;text-transform:uppercase;color:#64748b;border-bottom:1px solid #e2e8f0;white-space:nowrap}
.sm-table td{padding:8px 12px;border-bottom:1px solid #f1f5f9;font-size:12px;word-break:break-all;max-width:180px;vertical-align:top}
.sm-table tr:last-child td{border-bottom:none}
.sm-table td:first-child{font-weight:600;white-space:nowrap;word-break:normal;max-width:none}
.sm-table td a{color:#2563eb;font-size:11px}
.dash{color:#cbd5e1}
.empty-state{padding:60px 24px;text-align:center;color:#94a3b8}
.empty-state .icon{font-size:36px;margin-bottom:12px}
.empty-state p{font-size:14px}
.no-results{padding:40px;text-align:center;color:#94a3b8;font-size:14px}
.email-link{color:#2563eb;text-decoration:none;font-size:12px}
.tag{display:inline-block;background:#eff6ff;color:#1d4ed8;border:1px solid #bfdbfe;border-radius:4px;padding:2px 8px;font-size:11px;font-weight:700;margin-right:4px;white-space:nowrap;min-width:30px;text-align:center}
</style>
</head>
<body>
<aside id="sidebar">
  <div id="sidebar-header">
    <div class="logo">Delta Force</div>
    <div class="title">REG MKT Teammate</div>
  </div>
  <ul id="nav-list"></ul>
</aside>
<div id="main">
  <div id="toolbar">
    <h2 id="page-title">Index</h2>
    <input id="search" type="search" placeholder="Search all content…" autocomplete="off">
  </div>
  <div id="content"></div>
</div>

<script>
const RAW = """ + DATA_JSON + r""";

// ── PIC badge helper ──────────────────────────────────────────
function picClass(pic){
  const p = (pic||'').toLowerCase();
  if(p.includes(',') || (p.includes('tim') && p.length > 5)) return 'pic-multi';
  if(p.includes('tim'))    return 'pic-tim';
  if(p.includes('ruru'))   return 'pic-ruru';
  if(p.includes('jeremy')) return 'pic-jeremy';
  if(p.includes('hua'))    return 'pic-hua';
  if(p.includes('mye'))    return 'pic-mye';
  return 'pic-multi';
}

// ── Link detector ─────────────────────────────────────────────
function linkify(text){
  if(!text) return '';
  const esc = text.replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
  return esc.replace(/(https?:\/\/[^\s]+)/g, '<a href="$1" target="_blank" rel="noopener">$1</a>');
}

// ── Render helpers ────────────────────────────────────────────
function field(label, value){
  if(!value || value === '—') return '';
  return `<div class="field">
    <div class="field-label">${label}</div>
    <div class="field-value">${linkify(value)}</div>
  </div>`;
}

function renderCards(rows, query){
  if(!rows || rows.length === 0)
    return `<div class="empty-state"><div class="icon">📭</div><p>No content yet — check back later.</p></div>`;
  const q = (query||'').toLowerCase();
  const filtered = rows.filter(r => {
    if(!q) return true;
    return Object.values(r).some(v => (v||'').toLowerCase().includes(q));
  });
  if(filtered.length === 0)
    return `<div class="no-results">No results for "<strong>${q}</strong>"</div>`;
  return `<div class="cards-grid">${filtered.map(r => {
    const hasLinks = r.links && r.links !== '—';
    return `<div class="card">
      <div class="card-header" onclick="toggleCard(this)">
        <div class="card-scenario">${r.scenario||'(Untitled)'}</div>
        <span class="card-pic ${picClass(r.pic)}">${r.pic||''}</span>
        <span class="card-chevron">▼</span>
      </div>
      <div class="card-body">
        ${field('Process Steps', r.steps)}
        ${field('What to Prepare', r.prepare)}
        ${field('Timeline', r.timeline)}
        ${field('Common Q&A', r.qa)}
        ${hasLinks ? `<div class="field"><div class="field-label">Related Links</div><div class="field-value">${r.links_url ? `<a href="${r.links_url}" target="_blank" rel="noopener" class="field-link">${r.links}</a>` : linkify(r.links)}</div></div>` : ''}
      </div>
    </div>`;
  }).join('')}</div>`;
}

function tabIdFromName(name){
  const m={
    'Contactor':'_contactor','SM Links':'_sm',
    'Budget & Plan':'tab1','OM':'tab2','Assets':'tab3',
    'Social Media':'tab4','PR':'tab5','Esports':'tab7',
    'Cross-Industry & IP Partnerships':'tab8','Virtual Item Rewards':'tab9',
    'CCP':'tab10','Website':'tab11','Patch Updates':'tab12','Store':'tab13'
  };
  return m[name]||null;
}

function renderIndex(rows){
  return `<p class="page-title">Tab Index</p>
  <table class="data-table">
    <thead><tr><th style="width:48px">Tab</th><th style="width:180px">Name</th><th>Description</th></tr></thead>
    <tbody>${rows.map(r=>{
      const tid=tabIdFromName(r.name);
      const nameEl=tid
        ? `<a href="#" onclick="navigate(\'${tid}\');return false;" style="color:#2563eb;font-weight:600;text-decoration:none">${r.name} ↗</a>`
        : `<span style="font-weight:600">${r.name}</span>`;
      return `<tr>
        <td><span class="tag">${r.tab}</span></td>
        <td>${nameEl}</td>
        <td style="color:#475569">${r.desc}</td>
      </tr>`;
    }).join('')}</tbody>
  </table>`;
}

function renderContactor(data){
  const teamRows = data.team.map(r=>`<tr>
    <td style="font-weight:700">${r.name}</td>
    <td class="resp-col">${r.resp}</td>
    <td><a href="mailto:${r.email}" class="email-link">${r.email}</a></td>
  </tr>`).join('');
  const locRows = data.contactors.map(r=>`<tr>
    <td style="font-weight:700">${r.region}</td>
    <td>${r.contact}</td>
    <td><a href="mailto:${r.email}" class="email-link">${r.email}</a></td>
  </tr>`).join('');
  return `<p class="page-title">Contactor</p>
  <div class="section-title">Regional Team</div>
  <table class="data-table">
    <thead><tr><th>Member</th><th>Responsibilities</th><th>Email</th></tr></thead>
    <tbody>${teamRows}</tbody>
  </table>
  <div class="section-title">Local MKT Contactors</div>
  <table class="data-table">
    <thead><tr><th>Region</th><th>MKT Contactor</th><th>Email</th></tr></thead>
    <tbody>${locRows}</tbody>
  </table>`;
}

function smCell(val){
  if(!val || val==='—') return `<span class="dash">—</span>`;
  if(val.startsWith('http')) return `<a href="${val}" target="_blank" rel="noopener">↗ Link</a>`;
  return val;
}

function renderSM(sm){
  const offHead = sm.off_hdr.map(h=>`<th>${h}</th>`).join('');
  const offBody = sm.off.map(r=>`<tr>${r.map(v=>`<td>${smCell(v)}</td>`).join('')}</tr>`).join('');
  const nonHead = sm.non_hdr.map(h=>`<th>${h}</th>`).join('');
  const nonBody = sm.non.map(r=>`<tr>${r.map(v=>`<td>${smCell(v)}</td>`).join('')}</tr>`).join('');
  const espHead = sm.esp_hdr && sm.esp_hdr.length ? `<thead><tr>${sm.esp_hdr.map(h=>`<th>${h}</th>`).join('')}</tr></thead>` : '';
  const espSection = sm.esp.length ? `
    <div class="section-title">Esports Channels</div>
    <div class="sm-table-wrap">
      <table class="sm-table">${espHead}<tbody>${sm.esp.map(r=>`<tr>${r.map(v=>`<td>${smCell(v)}</td>`).join('')}</tr>`).join('')}</tbody>
      </table>
    </div>` : '';
  return `<p class="page-title">Social Media Links</p>
  <div class="section-title">Official Channels</div>
  <div class="sm-table-wrap">
    <table class="sm-table"><thead><tr>${offHead}</tr></thead><tbody>${offBody}</tbody></table>
  </div>
  <div class="section-title">Non-Official Channels</div>
  <div class="sm-table-wrap">
    <table class="sm-table"><thead><tr>${nonHead}</tr></thead><tbody>${nonBody}</tbody></table>
  </div>
  ${espSection}`;
}

// ── Navigation ────────────────────────────────────────────────
let activeId = '_index';

function getLabel(id){
  const item = RAW.nav.find(n=>n.id===id);
  return item ? item.label : id;
}

function buildNav(){
  const ul = document.getElementById('nav-list');
  RAW.nav.forEach(item => {
    const li = document.createElement('li');
    if(item.sep){ li.className='sep'; ul.appendChild(li); return; }
    li.className = 'nav-item';
    li.innerHTML = `<a href="#" data-id="${item.id}">${item.label}</a>`;
    li.querySelector('a').addEventListener('click', e => {
      e.preventDefault();
      navigate(item.id);
    });
    ul.appendChild(li);
  });
}

function navigate(id){
  activeId = id;
  document.querySelectorAll('#nav-list a').forEach(a => {
    a.classList.toggle('active', a.dataset.id === id);
  });
  document.getElementById('page-title').textContent = getLabel(id);
  document.getElementById('search').value = '';
  renderContent('');
}

function renderContent(query){
  const el = document.getElementById('content');
  const id  = activeId;
  if(id === '_index')     { el.innerHTML = renderIndex(RAW.tabs['_index']); return; }
  if(id === '_contactor') { el.innerHTML = renderContactor(RAW.tabs['_contactor']); return; }
  if(id === '_sm')        { el.innerHTML = renderSM(RAW.tabs['_sm']); return; }
  el.innerHTML = renderCards(RAW.tabs[id] || [], query);
}

// ── Search ────────────────────────────────────────────────────
let searchTimer;
document.getElementById('search').addEventListener('input', e => {
  clearTimeout(searchTimer);
  searchTimer = setTimeout(() => renderContent(e.target.value), 120);
});

// ── Card toggle ───────────────────────────────────────────────
function toggleCard(header){
  const body    = header.nextElementSibling;
  const chevron = header.querySelector('.card-chevron');
  body.classList.toggle('open');
  chevron.classList.toggle('open');
}

// ── Init ──────────────────────────────────────────────────────
buildNav();
navigate('_index');
</script>
</body>
</html>"""

out_html = os.path.join(HERE, 'index.html')
with open(out_html, 'w', encoding='utf-8') as f:
    f.write(HTML)

print(f"Generated: {out_html}")
